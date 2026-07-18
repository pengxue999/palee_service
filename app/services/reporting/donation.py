import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session, joinedload

from app.models.donation_category import DonationCategory
from app.models.donation import Donation
from app.models.donor import Donor
from app.utils.donation_category import is_cash_donation_name
from app.services.reporting.common import (
    create_csv_writer,
    current_timestamp,
    finalize_binary_export,
    finalize_csv_export,
)


def get_donation_report(
    db: Session,
    *,
    donor_id: Optional[str] = None,
    donation_category: Optional[str] = None,
    year: Optional[int] = None,
) -> dict[str, Any]:
    donor_name = None
    if donor_id:
        donor = db.query(Donor).filter(Donor.donor_id == donor_id).first()
        if donor:
            donor_name = f"{donor.donor_name} {donor.donor_lastname}".strip()

    selected_category_name = (
        (donation_category or "").strip()
        if donation_category
        else None
    )

    query = db.query(Donation).options(
        joinedload(Donation.donor),
        joinedload(Donation.donation_category),
    )

    if donor_id:
        query = query.filter(Donation.donor_id == donor_id)
    if selected_category_name:
        query = query.join(Donation.donation_category).filter(
            DonationCategory.donation_category_name == selected_category_name
        )
    if year:
        query = query.filter(Donation.donation_date.isnot(None))
        query = query.filter(Donation.donation_date >= datetime(year, 1, 1).date())
        query = query.filter(Donation.donation_date <= datetime(year, 12, 31).date())

    donations = query.order_by(Donation.donation_date.desc(), Donation.donation_id.desc()).all()

    donation_rows: list[dict[str, Any]] = []
    total_amount = Decimal("0")
    cash_donation_count = 0
    category_summary: dict[str, int] = {}

    for donation in donations:
        amount = Decimal(donation.amount or 0)
        category_name = (
            donation.donation_category.donation_category_name
            if donation.donation_category is not None
            else "-"
        )
        is_cash = is_cash_donation_name(category_name)

        if is_cash:
            total_amount += amount
            cash_donation_count += 1
        category_summary[category_name] = category_summary.get(category_name, 0) + 1

        donation_rows.append(
            {
                "donation_id": donation.donation_id,
                "donor_name": (
                    f"{donation.donor.donor_name} {donation.donor.donor_lastname}".strip()
                    if donation.donor
                    else "-"
                ),
                "category": category_name,
                "donation_name": donation.donation_name,
                "amount": float(amount),
                "amount_display": f"{int(amount):,}" if is_cash else f"{float(amount):,.0f}",
                "unit": donation.unit or "-",
                "donation_date": donation.donation_date.strftime("%d-%m-%Y")
                if donation.donation_date
                else "-",
                "is_cash": is_cash,
            }
        )

    return {
        "filters": {
            "donor_id": donor_id,
            "donor_name": donor_name,
            "donation_category": selected_category_name,
            "donation_category_name": selected_category_name,
            "year": year,
        },
        "summary": {
            "total_count": len(donation_rows),
            "cash_donation_count": cash_donation_count,
            "category_count": len(category_summary),
            "total_amount": float(total_amount),
            "categories": category_summary,
        },
        "donations": donation_rows,
    }


def export_donation_report(
    db: Session,
    *,
    donor_id: Optional[str] = None,
    donation_category: Optional[str] = None,
    year: Optional[int] = None,
    format: str = "excel",
) -> dict[str, Any]:
    report_data = get_donation_report(
        db,
        donor_id=donor_id,
        donation_category=donation_category,
        year=year,
    )
    donations = report_data["donations"]
    normalized_format = format.lower()

    if normalized_format == "csv":
        output, writer = create_csv_writer()
        writer.writerow(
            [
                "ລະຫັດ",
                "ຜູ້ບໍລິຈາກ",
                "ປະເພດບໍລິຈາກ",
                "ລາຍການບໍລິຈາກ",
                "ຈຳນວນ / ມູນຄ່າ",
                "ຫົວໜ່ວຍ",
                "ວັນທີບໍລິຈາກ",
            ]
        )
        for item in donations:
            writer.writerow(
                [
                    item["donation_id"],
                    item["donor_name"],
                    item["category"],
                    item["donation_name"],
                    item["amount_display"],
                    item["unit"],
                    item["donation_date"],
                ]
            )

        filename = f"donation_report_{current_timestamp()}.csv"
        return finalize_csv_export(output, filename=filename, total_records=len(donations))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Donations"

    title_fill = PatternFill("solid", fgColor="1D4ED8")
    summary_fill = PatternFill("solid", fgColor="EFF6FF")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = "ລາຍງານການບໍລິຈາກ"
    title_cell.font = Font(name="Noto Sans Lao", size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    summary_rows = [
        ("ວັນທີສ້າງ", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("ຈຳນວນລາຍການ", report_data["summary"]["total_count"]),
        ("ສະຖິຕິຕາມປະເພດ", report_data["summary"]["category_count"]),
        ("ຍອດລວມມູນຄ່າເງິນສົດ", f"{int(report_data['summary']['total_amount']):,} ₭"),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        for column in range(1, 3):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = summary_fill
            cell.border = thin_border
            cell.font = Font(name="Noto Sans Lao", bold=column == 1)

    header_row = 8
    headers = [
        "ລະຫັດ",
        "ຜູ້ບໍລິຈາກ",
        "ປະເພດບໍລິຈາກ",
        "ລາຍການບໍລິຈາກ",
        "ຈຳນວນ / ມູນຄ່າ",
        "ຫົວໜ່ວຍ",
        "ວັນທີບໍລິຈາກ",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(name="Noto Sans Lao", bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, item in enumerate(donations, start=header_row + 1):
        values = [
            item["donation_id"],
            item["donor_name"],
            item["category"],
            item["donation_name"],
            item["amount_display"],
            item["unit"],
            item["donation_date"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Noto Sans Lao")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:G{max(header_row, header_row + len(donations))}"
    for column_name, width in {
        "A": 12,
        "B": 26,
        "C": 18,
        "D": 28,
        "E": 16,
        "F": 14,
        "G": 14,
    }.items():
        sheet.column_dimensions[column_name].width = width

    output = io.BytesIO()
    workbook.save(output)
    filename = f"donation_report_{current_timestamp()}.xlsx"
    return finalize_binary_export(
        output.getvalue(),
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        total_records=len(donations),
    )