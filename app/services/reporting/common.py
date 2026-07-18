import base64
import csv
import io
from datetime import datetime
from typing import Iterable, Mapping, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session

from app.models.academic_years import AcademicYear
from app.models.district import District
from app.models.province import Province
from app.models.teacher import Teacher


REPORT_FONT_NAME = "Noto Sans Lao"


def create_excel_theme(
    *,
    title_color: str = "1D4ED8",
    header_color: str = "DBEAFE",
    summary_color: str = "EFF6FF",
) -> dict[str, object]:
    return {
        "title_fill": PatternFill("solid", fgColor=title_color),
        "header_fill": PatternFill("solid", fgColor=header_color),
        "summary_fill": PatternFill("solid", fgColor=summary_color),
        "border": Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        ),
    }


def apply_excel_title(
    sheet,
    *,
    title: str,
    from_column: int,
    to_column: int,
    row: int = 1,
    theme: Optional[Mapping[str, object]] = None,
) -> None:
    excel_theme = theme or create_excel_theme()
    start_cell = f"{get_column_letter(from_column)}{row}"
    end_cell = f"{get_column_letter(to_column)}{row}"
    sheet.merge_cells(f"{start_cell}:{end_cell}")
    cell = sheet[start_cell]
    cell.value = title
    cell.font = Font(name=REPORT_FONT_NAME, size=16, bold=True, color="FFFFFF")
    cell.fill = excel_theme["title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 26


def write_excel_key_value_rows(
    sheet,
    *,
    rows: Sequence[tuple[str, object]],
    start_row: int,
    label_column: int = 1,
    value_column: int = 2,
    theme: Optional[Mapping[str, object]] = None,
) -> None:
    excel_theme = theme or create_excel_theme()
    for row_index, (label, value) in enumerate(rows, start=start_row):
        sheet.cell(row=row_index, column=label_column, value=label)
        sheet.cell(row=row_index, column=value_column, value=value)
        for column in range(label_column, value_column + 1):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = excel_theme["summary_fill"]
            cell.border = excel_theme["border"]
            cell.font = Font(name=REPORT_FONT_NAME, bold=column == label_column)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_excel_table_headers(
    sheet,
    *,
    headers: Sequence[str],
    row_index: int,
    theme: Optional[Mapping[str, object]] = None,
) -> None:
    excel_theme = theme or create_excel_theme()
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=header)
        cell.font = Font(name=REPORT_FONT_NAME, bold=True)
        cell.fill = excel_theme["header_fill"]
        cell.border = excel_theme["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel_table_rows(
    sheet,
    *,
    rows: Iterable[Sequence[object]],
    start_row: int,
    theme: Optional[Mapping[str, object]] = None,
) -> int:
    excel_theme = theme or create_excel_theme()
    current_row = start_row
    for values in rows:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current_row, column=column_index, value=value)
            cell.font = Font(name=REPORT_FONT_NAME)
            cell.border = excel_theme["border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        current_row += 1
    return current_row - 1


def set_excel_column_widths(sheet, widths: Mapping[int | str, float]) -> None:
    for column_ref, width in widths.items():
        column_name = (
            get_column_letter(column_ref)
            if isinstance(column_ref, int)
            else column_ref
        )
        sheet.column_dimensions[column_name].width = width


def finalize_workbook_export(
    workbook: Workbook,
    *,
    filename: str,
    total_records: int,
) -> dict[str, object]:
    output = io.BytesIO()
    workbook.save(output)
    return finalize_binary_export(
        output.getvalue(),
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        total_records=total_records,
    )


def format_report_date(value: Optional[str]) -> str:
    if not value:
        return "-"
    for date_format in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, date_format).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return value


def format_report_datetime(value: Optional[datetime] = None) -> str:
    return (value or datetime.now()).strftime("%d-%m-%Y %H:%M")


def format_report_currency(value: object, suffix: str = "₭") -> str:
    try:
        numeric = float(value or 0)
    except (TypeError, ValueError):
        numeric = 0
    return f"{numeric:,.0f} {suffix}".strip()


def create_csv_writer() -> tuple[io.StringIO, csv.writer]:
    output = io.StringIO()
    output.write("\ufeff")
    return output, csv.writer(output)


def finalize_csv_export(
    output: io.StringIO,
    *,
    filename: str,
    total_records: int,
) -> dict[str, object]:
    csv_content = output.getvalue()
    output.close()
    return {
        "filename": filename,
        "content_type": "text/csv; charset=utf-8-sig",
        "data": csv_content,
        "total_records": total_records,
    }


def finalize_binary_export(
    payload: bytes,
    *,
    filename: str,
    content_type: str,
    total_records: int,
) -> dict[str, object]:
    return {
        "filename": filename,
        "content_type": content_type,
        "data": base64.b64encode(payload).decode("ascii"),
        "total_records": total_records,
    }


def current_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def resolve_academic_year_name(db: Session, academic_id: Optional[str]) -> Optional[str]:
    if not academic_id:
        return None

    academic = (
        db.query(AcademicYear)
        .filter(AcademicYear.academic_id == academic_id)
        .first()
    )
    if academic:
        return academic.academic_year
    return None


def resolve_province_name(db: Session, province_id: Optional[int]) -> Optional[str]:
    if not province_id:
        return None

    province = db.query(Province).filter(Province.province_id == province_id).first()
    if province:
        return province.province_name
    return None


def resolve_district_name(db: Session, district_id: Optional[int]) -> Optional[str]:
    if not district_id:
        return None

    district = db.query(District).filter(District.district_id == district_id).first()
    if district:
        return district.district_name
    return None


def resolve_teacher_name(db: Session, teacher_id: Optional[str]) -> Optional[str]:
    if not teacher_id:
        return None

    teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
    if teacher:
        return f"{teacher.teacher_name} {teacher.teacher_lastname}"
    return None