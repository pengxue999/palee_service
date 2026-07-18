from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook

from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from app.models.expense import Expense
from app.models.expense_category import ExpenseCategory
from app.models.income import Income
from app.services.reporting.common import (
    apply_excel_title,
    create_csv_writer,
    create_excel_theme,
    current_timestamp,
    finalize_workbook_export,
    finalize_csv_export,
    format_report_currency,
    format_report_date,
    format_report_datetime,
    set_excel_column_widths,
    write_excel_key_value_rows,
    write_excel_table_headers,
    write_excel_table_rows,
    resolve_academic_year_name,
)


def get_finance_report(
    db: Session,
    academic_id: Optional[str] = None,
    year: Optional[int] = None,
) -> Dict[str, Any]:
    income_query = db.query(Income)
    expense_query = db.query(Expense).join(ExpenseCategory)

    if year:
        income_query = income_query.filter(extract("year", Income.income_date) == year)
        expense_query = expense_query.filter(
            extract("year", Expense.expense_date) == year
        )

    incomes = income_query.all()
    expenses = expense_query.all()

    total_income = sum(float(income.amount) for income in incomes)
    total_expense = sum(float(expense.amount) for expense in expenses)
    balance = total_income - total_expense

    tuition_income = sum(
        float(income.amount) for income in incomes if income.tuition_payment_id is not None
    )
    donation_income = sum(
        float(income.amount) for income in incomes if income.donation_id is not None
    )
    other_income = total_income - tuition_income - donation_income

    income_breakdown = []
    if total_income > 0:
        income_breakdown = [
            {
                "category": "ຄ່າຮຽນ",
                "amount": tuition_income,
                "percentage": round(tuition_income / total_income * 100, 2),
            },
            {
                "category": "ການບໍລິຈາກ",
                "amount": donation_income,
                "percentage": round(donation_income / total_income * 100, 2),
            },
            {
                "category": "ອື່ນໆ",
                "amount": other_income,
                "percentage": round(other_income / total_income * 100, 2),
            },
        ]

    expense_breakdown = []
    for category in db.query(ExpenseCategory).all():
        category_expense = sum(
            float(expense.amount)
            for expense in expenses
            if expense.expense_category_id == category.expense_category_id
        )
        if category_expense > 0:
            percentage = (
                round(category_expense / total_expense * 100, 2)
                if total_expense > 0
                else 0
            )
            expense_breakdown.append(
                {
                    "category": category.expense_category,
                    "amount": category_expense,
                    "percentage": percentage,
                }
            )

    expense_breakdown.sort(key=lambda item: item["amount"], reverse=True)

    years_with_data = []
    min_income_year = db.query(func.min(extract("year", Income.income_date))).scalar()
    max_income_year = db.query(func.max(extract("year", Income.income_date))).scalar()
    min_expense_year = db.query(func.min(extract("year", Expense.expense_date))).scalar()
    max_expense_year = db.query(func.max(extract("year", Expense.expense_date))).scalar()

    if min_income_year is not None:
        years_with_data.append(int(min_income_year))
    if max_income_year is not None:
        years_with_data.append(int(max_income_year))
    if min_expense_year is not None:
        years_with_data.append(int(min_expense_year))
    if max_expense_year is not None:
        years_with_data.append(int(max_expense_year))

    if years_with_data:
        start_year = min(years_with_data)
        end_year = max(years_with_data)
    else:
        start_year = datetime.now().year
        end_year = start_year

    yearly_data = []
    for current_year in range(start_year, end_year + 1):
        year_income = (
            db.query(func.sum(Income.amount))
            .filter(extract("year", Income.income_date) == current_year)
            .scalar()
            or 0
        )
        year_expense = (
            db.query(func.sum(Expense.amount))
            .filter(extract("year", Expense.expense_date) == current_year)
            .scalar()
            or 0
        )
        yearly_data.append(
            {
                "year": current_year,
                "income": float(year_income),
                "expense": float(year_expense),
                "balance": float(year_income) - float(year_expense),
            }
        )

    income_list = [
        {
            "income_id": income.income_id,
            "amount": float(income.amount),
            "description": income.description,
            "income_date": (
                income.income_date.strftime("%Y-%m-%d") if income.income_date else None
            ),
            "source": (
                "ຄ່າຮຽນ"
                if income.tuition_payment_id
                else "ການບໍລິຈາກ"
                if income.donation_id
                else "ອື່ນໆ"
            ),
        }
        for income in incomes
    ]

    expense_list = [
        {
            "expense_id": expense.expense_id,
            "amount": float(expense.amount),
            "description": expense.description,
            "expense_date": (
                expense.expense_date.strftime("%Y-%m-%d")
                if expense.expense_date
                else None
            ),
            "category": expense.category.expense_category if expense.category else "-",
        }
        for expense in expenses
    ]

    return {
        "filters": {
            "academic_id": academic_id,
            "academic_year_name": resolve_academic_year_name(db, academic_id),
            "year": year,
        },
        "summary": {
            "total_income": total_income,
            "total_expense": total_expense,
            "balance": balance,
        },
        "income_breakdown": income_breakdown,
        "expense_breakdown": expense_breakdown,
        "yearly_comparison": yearly_data,
        "incomes": income_list,
        "expenses": expense_list,
        "total_income_count": len(income_list),
        "total_expense_count": len(expense_list),
    }


def export_finance_report(
    db: Session,
    academic_id: Optional[str] = None,
    year: Optional[int] = None,
    tab: str = "overview",
    format: str = "excel",
) -> Dict[str, Any]:
    report_data = get_finance_report(db, academic_id=academic_id, year=year)
    incomes = report_data["incomes"]
    expenses = report_data["expenses"]
    normalized_tab = tab if tab in {"overview", "income", "expense"} else "overview"
    normalized_format = format.lower()

    def format_export_date(value: Optional[str]) -> str:
        if not value:
            return "-"
        try:
            return datetime.fromisoformat(value).strftime("%d-%m-%Y")
        except ValueError:
            return value

    def format_export_amount(value: object) -> str:
        try:
            numeric = float(value or 0)
        except (TypeError, ValueError):
            numeric = 0
        return f"{numeric:,.0f} ₭"

    if normalized_format == "csv":
        output, writer = create_csv_writer()

        if normalized_tab == "income":
            writer.writerow(["ລາຍການລາຍຮັບ"])
            writer.writerow(["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ແຫຼ່ງທີ່ມາ", "ວັນທີ"])
            for index, income in enumerate(incomes, start=1):
                writer.writerow(
                    [
                        index,
                        format_export_amount(income["amount"]),
                        income["description"] or "-",
                        income["source"],
                        format_export_date(income["income_date"]),
                    ]
                )
        elif normalized_tab == "expense":
            writer.writerow(["ລາຍການລາຍຈ່າຍ"])
            writer.writerow(["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ປະເພດ", "ວັນທີ"])
            for index, expense in enumerate(expenses, start=1):
                writer.writerow(
                    [
                        index,
                        format_export_amount(expense["amount"]),
                        expense["description"] or "-",
                        expense["category"],
                        format_export_date(expense["expense_date"]),
                    ]
                )
        else:
            writer.writerow(["ສະຫຼຸບການເງິນ"])
            writer.writerow(["ລາຍຮັບທັງໝົດ", report_data["summary"]["total_income"]])
            writer.writerow(["ລາຍຈ່າຍທັງໝົດ", report_data["summary"]["total_expense"]])
            writer.writerow(["ຍອດເຫຼືອ", report_data["summary"]["balance"]])
            writer.writerow([])
            writer.writerow(["ລາຍຮັບແຍກຕາມແຫຼ່ງ"])
            writer.writerow(["ແຫຼ່ງລາຍຮັບ", "ຈຳນວນເງິນ", "ເປີເຊັນ"])
            for item in report_data["income_breakdown"]:
                writer.writerow([item["category"], item["amount"], f"{item['percentage']}%"])
            writer.writerow([])
            writer.writerow(["ລາຍຈ່າຍແຍກຕາມປະເພດ"])
            writer.writerow(["ປະເພດລາຍຈ່າຍ", "ຈຳນວນເງິນ", "ເປີເຊັນ"])
            for item in report_data["expense_breakdown"]:
                writer.writerow([item["category"], item["amount"], f"{item['percentage']}%"])

        filters_desc = []
        if academic_id:
            filters_desc.append(f"academic_{academic_id}")
        if year:
            filters_desc.append(f"year_{year}")
        filters_desc.append(normalized_tab)

        filter_str = "_".join(filters_desc) if filters_desc else "ທັງໝົດ"
        filename = f"ລາຍງານການເງິນ_{filter_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        total_records = (
            len(incomes)
            if normalized_tab == "income"
            else len(expenses)
            if normalized_tab == "expense"
            else len(incomes) + len(expenses)
        )
        return finalize_csv_export(output, filename=filename, total_records=total_records)

    workbook = Workbook()
    theme = create_excel_theme(title_color="7C3AED", header_color="EDE9FE", summary_color="F5F3FF")

    if normalized_tab == "overview":
        summary_sheet = workbook.active
        summary_sheet.title = "Overview"
        income_sheet = workbook.create_sheet("Income Breakdown")
        expense_sheet = workbook.create_sheet("Expense Breakdown")

        apply_excel_title(
            summary_sheet,
            title="ສະຫຼຸບລາຍງານການເງິນ",
            from_column=1,
            to_column=3,
            theme=theme,
        )
        write_excel_key_value_rows(
            summary_sheet,
            rows=[
                ("ວັນທີສ້າງ", format_report_datetime()),
                ("ສົກຮຽນ", report_data["filters"].get("academic_year_name") or "ທັງໝົດ"),
                ("ປີ", report_data["filters"].get("year") or "ທັງໝົດ"),
                ("ລາຍຮັບທັງໝົດ", format_report_currency(report_data["summary"]["total_income"])),
                ("ລາຍຈ່າຍທັງໝົດ", format_report_currency(report_data["summary"]["total_expense"])),
                ("ຍອດເຫຼືອ", format_report_currency(report_data["summary"]["balance"])),
            ],
            start_row=3,
            theme=theme,
        )

        income_header_row = 10
        write_excel_table_headers(
            summary_sheet,
            headers=["ແຫຼ່ງລາຍຮັບ", "ຈຳນວນເງິນ", "ເປີເຊັນ"],
            row_index=income_header_row,
            theme=theme,
        )
        write_excel_table_rows(
            summary_sheet,
            rows=[
                [item["category"], format_report_currency(item["amount"]), f"{item['percentage']}%"]
                for item in report_data["income_breakdown"]
            ],
            start_row=income_header_row + 1,
            theme=theme,
        )

        expense_header_row = max(income_header_row + 3, income_header_row + 1 + len(report_data["income_breakdown"]) + 2)
        write_excel_table_headers(
            summary_sheet,
            headers=["ປະເພດລາຍຈ່າຍ", "ຈຳນວນເງິນ", "ເປີເຊັນ"],
            row_index=expense_header_row,
            theme=theme,
        )
        write_excel_table_rows(
            summary_sheet,
            rows=[
                [item["category"], format_report_currency(item["amount"]), f"{item['percentage']}%"]
                for item in report_data["expense_breakdown"]
            ],
            start_row=expense_header_row + 1,
            theme=theme,
        )

        apply_excel_title(
            income_sheet,
            title="ລາຍການລາຍຮັບ",
            from_column=1,
            to_column=5,
            theme=theme,
        )
        write_excel_table_headers(
            income_sheet,
            headers=["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ແຫຼ່ງທີ່ມາ", "ວັນທີ"],
            row_index=3,
            theme=theme,
        )
        write_excel_table_rows(
            income_sheet,
            rows=[
                [
                    index,
                    format_report_currency(income["amount"]),
                    income["description"] or "-",
                    income["source"],
                    format_report_date(income["income_date"]),
                ]
                for index, income in enumerate(incomes, start=1)
            ],
            start_row=4,
            theme=theme,
        )

        apply_excel_title(
            expense_sheet,
            title="ລາຍການລາຍຈ່າຍ",
            from_column=1,
            to_column=5,
            theme=theme,
        )
        write_excel_table_headers(
            expense_sheet,
            headers=["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ປະເພດ", "ວັນທີ"],
            row_index=3,
            theme=theme,
        )
        write_excel_table_rows(
            expense_sheet,
            rows=[
                [
                    index,
                    format_report_currency(expense["amount"]),
                    expense["description"] or "-",
                    expense["category"],
                    format_report_date(expense["expense_date"]),
                ]
                for index, expense in enumerate(expenses, start=1)
            ],
            start_row=4,
            theme=theme,
        )

        for sheet in (summary_sheet, income_sheet, expense_sheet):
            sheet.freeze_panes = "A4" if sheet.title != "Overview" else "A11"
        set_excel_column_widths(summary_sheet, {1: 24, 2: 20, 3: 14})
        set_excel_column_widths(income_sheet, {1: 10, 2: 18, 3: 34, 4: 18, 5: 14})
        set_excel_column_widths(expense_sheet, {1: 10, 2: 18, 3: 34, 4: 18, 5: 14})
    else:
        sheet = workbook.active
        sheet.title = "Income" if normalized_tab == "income" else "Expense"
        title = "ລາຍການລາຍຮັບ" if normalized_tab == "income" else "ລາຍການລາຍຈ່າຍ"
        headers = ["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ແຫຼ່ງທີ່ມາ", "ວັນທີ"] if normalized_tab == "income" else ["ລຳດັບ", "ຈຳນວນເງິນ", "ລາຍລະອຽດ", "ປະເພດ", "ວັນທີ"]
        apply_excel_title(sheet, title=title, from_column=1, to_column=5, theme=theme)
        write_excel_key_value_rows(
            sheet,
            rows=[
                ("ວັນທີສ້າງ", format_report_datetime()),
                ("ສົກຮຽນ", report_data["filters"].get("academic_year_name") or "ທັງໝົດ"),
                ("ປີ", report_data["filters"].get("year") or "ທັງໝົດ"),
                ("ຈຳນວນລາຍການ", len(incomes) if normalized_tab == "income" else len(expenses)),
                ("ມູນຄ່າລວມ", format_report_currency(
                    report_data["summary"]["total_income"] if normalized_tab == "income" else report_data["summary"]["total_expense"]
                )),
            ],
            start_row=3,
            theme=theme,
        )
        write_excel_table_headers(sheet, headers=headers, row_index=9, theme=theme)
        write_excel_table_rows(
            sheet,
            rows=[
                [
                    index,
                    format_report_currency(item["amount"]),
                    item["description"] or "-",
                    item["source"] if normalized_tab == "income" else item["category"],
                    format_report_date(item["income_date"] if normalized_tab == "income" else item["expense_date"]),
                ]
                for index, item in enumerate(incomes if normalized_tab == "income" else expenses, start=1)
            ],
            start_row=10,
            theme=theme,
        )
        sheet.freeze_panes = "A10"
        sheet.auto_filter.ref = f"A9:E{max(9, 9 + len(incomes if normalized_tab == 'income' else expenses))}"
        set_excel_column_widths(sheet, {1: 10, 2: 18, 3: 34, 4: 18, 5: 14})

    total_records = (
        len(incomes)
        if normalized_tab == "income"
        else len(expenses)
        if normalized_tab == "expense"
        else len(incomes) + len(expenses)
    )
    filename = f"finance_report_{normalized_tab}_{current_timestamp()}.xlsx"
    return finalize_workbook_export(workbook, filename=filename, total_records=total_records)