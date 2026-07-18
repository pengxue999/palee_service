import io
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.fee import Fee
from app.models.level import Level
from app.models.registration import Registration
from app.models.registration_detail import RegistrationDetail
from app.models.student import Student
from app.models.subject import Subject
from app.models.subject_category import SubjectCategory
from app.models.subject_detail import SubjectDetail
from app.services.reporting.common import (
    create_csv_writer,
    current_timestamp,
    finalize_binary_export,
    finalize_csv_export,
    resolve_academic_year_name,
)


def get_popular_subject_level_detail_report(
    db: Session,
    *,
    subject_name: str,
    level_name: str,
    subject_category: Optional[str] = None,
    academic_id: Optional[str] = None,
) -> Dict[str, Any]:
    query = (
        db.query(
            Student.student_id,
            Student.student_name,
            Student.student_lastname,
            Student.school,
        )
        .select_from(RegistrationDetail)
        .join(Registration, RegistrationDetail.registration_id == Registration.registration_id)
        .join(Fee, RegistrationDetail.fee_id == Fee.fee_id)
        .join(SubjectDetail, Fee.subject_detail_id == SubjectDetail.subject_detail_id)
        .join(Subject, SubjectDetail.subject_id == Subject.subject_id)
        .join(
            SubjectCategory,
            Subject.subject_category_id == SubjectCategory.subject_category_id,
        )
        .join(Level, SubjectDetail.level_id == Level.level_id)
        .join(Student, Registration.student_id == Student.student_id)
        .filter(
            Subject.subject_name == subject_name,
            Level.level_name == level_name,
        )
    )

    if subject_category:
        query = query.filter(SubjectCategory.subject_category_name == subject_category)

    if academic_id:
        query = query.filter(Fee.academic_id == academic_id)

    students = [
        {
            "student_id": row.student_id,
            "student_name": row.student_name,
            "student_lastname": row.student_lastname,
            "full_name": f"{row.student_name} {row.student_lastname}",
            "school": row.school,
        }
        for row in query.group_by(
            Student.student_id,
            Student.student_name,
            Student.student_lastname,
            Student.school,
        )
        .order_by(Student.student_id.asc())
        .all()
    ]

    return {
        "filters": {
            "academic_id": academic_id,
            "academic_year_name": resolve_academic_year_name(db, academic_id),
            "subject_name": subject_name,
            "subject_category": subject_category,
            "level_name": level_name,
        },
        "summary": {
            "total_students": len(students),
        },
        "students": students,
    }


def export_popular_subject_level_detail_report(
    db: Session,
    *,
    subject_name: str,
    level_name: str,
    subject_category: Optional[str] = None,
    academic_id: Optional[str] = None,
    format: str = "excel",
) -> Dict[str, Any]:
    report_data = get_popular_subject_level_detail_report(
        db,
        subject_name=subject_name,
        level_name=level_name,
        subject_category=subject_category,
        academic_id=academic_id,
    )
    students = report_data["students"]
    normalized_format = format.lower()
    title = f"ລາຍຊື່ນັກຮຽນ {subject_name} {level_name}"

    if normalized_format == "csv":
        output, writer = create_csv_writer()
        writer.writerow([title])
        writer.writerow([
            "ໝວດ",
            report_data["filters"].get("subject_category") or "-",
        ])
        writer.writerow([
            "ສົກຮຽນ",
            report_data["filters"].get("academic_year_name") or "ທັງໝົດ",
        ])
        writer.writerow([])
        writer.writerow(["ລະຫັດ", "ຊື່", "ນາມສະກຸນ", "ມາຈາກໂຮງຮຽນ"])
        for student in students:
            writer.writerow(
                [
                    student["student_id"],
                    student["student_name"],
                    student["student_lastname"],
                    student["school"],
                ]
            )

        filename = f"popular_subject_level_{current_timestamp()}.csv"
        return finalize_csv_export(
            output,
            filename=filename,
            total_records=len(students),
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="CCFBF1")
    info_fill = PatternFill("solid", fgColor="F0FDFA")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    sheet.merge_cells("A1:D1")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Noto Sans Lao", size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    info_rows = [
        ("ສົກຮຽນ", report_data["filters"].get("academic_year_name") or "ທັງໝົດ"),
        ("ໝວດ", report_data["filters"].get("subject_category") or "-"),
        ("ຈຳນວນນັກຮຽນ", len(students)),
    ]
    for row_index, (label, value) in enumerate(info_rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        for column in range(1, 3):
            cell = sheet.cell(row=row_index, column=column)
            cell.font = Font(name="Noto Sans Lao", bold=column == 1)
            cell.fill = info_fill
            cell.border = thin_border

    headers = ["ລະຫັດ", "ຊື່", "ນາມສະກຸນ", "ມາຈາກໂຮງຮຽນ"]
    header_row = 7
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(name="Noto Sans Lao", bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, student in enumerate(students, start=header_row + 1):
        values = [
            student["student_id"],
            student["student_name"],
            student["student_lastname"],
            student["school"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Noto Sans Lao")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A8"
    for column_name, width in {"A": 18, "B": 18, "C": 20, "D": 36}.items():
        sheet.column_dimensions[column_name].width = width

    output = io.BytesIO()
    workbook.save(output)
    filename = f"popular_subject_level_{current_timestamp()}.xlsx"
    return finalize_binary_export(
        output.getvalue(),
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        total_records=len(students),
    )


def get_popular_subjects_report(
    db: Session,
    academic_id: Optional[str] = None,
) -> Dict[str, Any]:
    query = (
        db.query(
            Subject.subject_name,
            SubjectCategory.subject_category_name,
            Level.level_name,
            func.count(RegistrationDetail.registration_id).label("student_count"),
            func.count(func.distinct(RegistrationDetail.registration_id)).label(
                "unique_students"
            ),
            Fee.fee.label("fee_amount"),
        )
        .select_from(RegistrationDetail)
        .join(Registration, RegistrationDetail.registration_id == Registration.registration_id)
        .join(Fee, RegistrationDetail.fee_id == Fee.fee_id)
        .join(SubjectDetail, Fee.subject_detail_id == SubjectDetail.subject_detail_id)
        .join(Subject, SubjectDetail.subject_id == Subject.subject_id)
        .join(
            SubjectCategory,
            Subject.subject_category_id == SubjectCategory.subject_category_id,
        )
        .join(Level, SubjectDetail.level_id == Level.level_id)
    )

    if academic_id:
        query = query.filter(Fee.academic_id == academic_id)

    subject_stats = (
        query.group_by(Subject.subject_name, SubjectCategory.subject_category_name)
        .with_entities(
            Subject.subject_name,
            SubjectCategory.subject_category_name.label("subject_category"),
            func.count(func.distinct(RegistrationDetail.registration_id)).label(
                "student_count"
            ),
            func.count(SubjectDetail.level_id).label("level_count"),
            func.avg(Fee.fee).label("avg_fee"),
        )
        .all()
    )

    level_stats = (
        query.group_by(
            Subject.subject_name,
            SubjectCategory.subject_category_name,
            Level.level_name,
            Fee.fee,
        )
        .with_entities(
            Subject.subject_name,
            SubjectCategory.subject_category_name.label("subject_category"),
            Level.level_name,
            func.count(func.distinct(RegistrationDetail.registration_id)).label(
                "student_count"
            ),
            Fee.fee.label("fee_amount"),
        )
        .all()
    )

    total_students = (
        db.query(RegistrationDetail)
        .join(Fee, RegistrationDetail.fee_id == Fee.fee_id)
        .filter(Fee.academic_id == academic_id if academic_id else True)
        .distinct(RegistrationDetail.registration_id)
        .count()
    )

    subjects_data = []
    for stat in subject_stats:
        percentage = (stat.student_count / total_students * 100) if total_students > 0 else 0
        subjects_data.append(
            {
                "subject_name": stat.subject_name,
                "subject_category": stat.subject_category,
                "student_count": stat.student_count,
                "level_count": stat.level_count,
                "avg_fee": float(stat.avg_fee) if stat.avg_fee else 0,
                "percentage": round(percentage, 2),
            }
        )

    subjects_data.sort(key=lambda item: item["student_count"], reverse=True)

    level_data = [
        {
            "subject_name": stat.subject_name,
            "subject_category": stat.subject_category,
            "level_name": stat.level_name,
            "student_count": stat.student_count,
            "fee_amount": float(stat.fee_amount) if stat.fee_amount else 0,
        }
        for stat in level_stats
    ]
    level_data.sort(key=lambda item: (item["subject_name"], item["level_name"]))

    category_stats = {}
    for subject in subjects_data:
        category = subject["subject_category"]
        if category not in category_stats:
            category_stats[category] = 0
        category_stats[category] += subject["student_count"]

    return {
        "filters": {
            "academic_id": academic_id,
            "academic_year_name": resolve_academic_year_name(db, academic_id),
        },
        "summary": {
            "total_students": total_students,
            "total_subjects": len(subjects_data),
            "total_categories": len(category_stats),
        },
        "subjects": subjects_data,
        "levels": level_data,
        "categories": category_stats,
    }


def export_popular_subjects_report(
    db: Session,
    academic_id: Optional[str] = None,
    format: str = "excel",
) -> Dict[str, Any]:
    report_data = get_popular_subjects_report(db, academic_id=academic_id)
    subjects = report_data["subjects"]
    levels = sorted(
        report_data["levels"],
        key=lambda item: (
            -int(item.get("student_count") or 0),
            item.get("subject_name") or "",
            item.get("level_name") or "",
        ),
    )
    summary = report_data["summary"]
    normalized_format = format.lower()

    if normalized_format == "csv":
        output, writer = create_csv_writer()
        writer.writerow(["ລາຍງານວິຊາຍອດນິຍົມ"])
        writer.writerow([])
        writer.writerow(["ວິຊາ", "ໝວດ", "ນັກຮຽນ", "ລະດັບ", "ສັດສ່ວນ"])
        for item in subjects:
            writer.writerow(
                [
                    item["subject_name"],
                    item["subject_category"],
                    item["student_count"],
                    item["level_count"],
                    f"{float(item['percentage']):.2f}%",
                ]
            )

        writer.writerow([])
        writer.writerow(["ລາຍລະອຽດຕາມລະດັບ/ຊັ້ນ"])
        writer.writerow(["ວິຊາ", "ໝວດ", "ລະດັບ/ຊັ້ນ", "ນັກຮຽນ"])
        for item in levels:
            writer.writerow(
                [
                    item["subject_name"],
                    item["subject_category"],
                    item["level_name"],
                    item["student_count"],
                ]
            )

        filename = f"popular_subjects_report_{current_timestamp()}.csv"
        return finalize_csv_export(output, filename=filename, total_records=len(subjects))

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    subjects_sheet = workbook.create_sheet("Subjects")
    levels_sheet = workbook.create_sheet("Levels")

    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="CCFBF1")
    summary_fill = PatternFill("solid", fgColor="F0FDFA")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def style_title(sheet, cell_range: str, title: str) -> None:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = title
        cell.font = Font(name="Noto Sans Lao", size=16, bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_header_row(sheet, row_index: int, headers: list[str]) -> None:
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=header)
            cell.font = Font(name="Noto Sans Lao", bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    style_title(summary_sheet, "A1:B1", "ສະຫຼຸບວິຊາຍອດນິຍົມ")
    summary_rows = [
        ("ວັນທີສ້າງ", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("ສົກຮຽນ", report_data["filters"].get("academic_year_name") or "ທັງໝົດ"),
        ("ຈຳນວນນັກຮຽນ", summary["total_students"]),
        ("ຈຳນວນວິຊາ", summary["total_subjects"]),
        ("ຈຳນວນໝວດວິຊາ", summary["total_categories"]),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)
        for column in range(1, 3):
            cell = summary_sheet.cell(row=row_index, column=column)
            cell.font = Font(name="Noto Sans Lao", bold=column == 1)
            cell.fill = summary_fill
            cell.border = thin_border

    category_start_row = len(summary_rows) + 5
    style_header_row(summary_sheet, category_start_row, ["ໝວດວິຊາ", "ຈຳນວນນັກຮຽນ"])
    for row_index, (category, count) in enumerate(report_data["categories"].items(), start=category_start_row + 1):
        summary_sheet.cell(row=row_index, column=1, value=category)
        summary_sheet.cell(row=row_index, column=2, value=count)
        for column in range(1, 3):
            cell = summary_sheet.cell(row=row_index, column=column)
            cell.font = Font(name="Noto Sans Lao")
            cell.border = thin_border

    style_title(subjects_sheet, "A1:E1", "ລາຍການວິຊານິຍົມ")
    subject_headers = ["ວິຊາ", "ໝວດ", "ນັກຮຽນ", "ລະດັບ", "ສັດສ່ວນ"]
    style_header_row(subjects_sheet, 3, subject_headers)
    for row_index, item in enumerate(subjects, start=4):
        values = [
            item["subject_name"],
            item["subject_category"],
            item["student_count"],
            item["level_count"],
            f"{float(item['percentage']):.2f}%",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = subjects_sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Noto Sans Lao")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    style_title(levels_sheet, "A1:D1", "ລາຍລະອຽດຕາມລະດັບ/ຊັ້ນ")
    level_headers = ["ວິຊາ", "ໝວດ", "ລະດັບ/ຊັ້ນ", "ນັກຮຽນ"]
    style_header_row(levels_sheet, 3, level_headers)
    for row_index, item in enumerate(levels, start=4):
        values = [
            item["subject_name"],
            item["subject_category"],
            item["level_name"],
            item["student_count"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = levels_sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Noto Sans Lao")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    for sheet in [summary_sheet, subjects_sheet, levels_sheet]:
        sheet.freeze_panes = "A3"

    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 24
    for column_name, width in {"A": 28, "B": 18, "C": 12, "D": 12, "E": 14}.items():
        subjects_sheet.column_dimensions[column_name].width = width
    for column_name, width in {"A": 28, "B": 18, "C": 18, "D": 12}.items():
        levels_sheet.column_dimensions[column_name].width = width

    output = io.BytesIO()
    workbook.save(output)
    filename = f"popular_subjects_report_{current_timestamp()}.xlsx"
    return finalize_binary_export(
        output.getvalue(),
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        total_records=len(subjects),
    )